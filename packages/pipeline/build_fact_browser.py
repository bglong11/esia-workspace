#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ESIA Table Fact Browser Generator

Transforms _meta.json table data into an interactive HTML viewer and Excel workbook.
Designed for ESIA reviewers (IFC, World Bank, ADB compliance checking).

Key Features:
- Interactive HTML with global search and collapsible sections
- Excel workbook organized by ESIA category
- Page provenance prominently displayed for every table
- No external dependencies (pure Python + openpyxl)

Usage:
    python build_fact_browser.py --input document_meta.json --output ./output_dir
    python build_fact_browser.py -i data/outputs/ESIA_Report_meta.json -o data/outputs/

Author: ESIA Workspace Pipeline
Date: December 2025
"""

import json
import argparse
import re
import html
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime

# Optional: openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will be disabled.")
    print("Install with: pip install openpyxl")


# =============================================================================
# CSS Styles (Inline)
# =============================================================================
CSS_STYLES = """
:root {
    --primary: #1f4e79;
    --secondary: #2d7d9a;
    --accent: #e8f4f8;
    --accent-dark: #c5e3ed;
    --text: #333;
    --text-light: #666;
    --border: #ddd;
    --background: #f5f5f5;
    --white: #ffffff;
    --success: #28a745;
    --warning: #ffc107;
    --danger: #dc3545;
}

* {
    box-sizing: border-box;
    margin: 0;
    padding: 0;
}

body {
    font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, 'Roboto', sans-serif;
    line-height: 1.6;
    color: var(--text);
    background: var(--background);
}

/* Header */
header {
    background: linear-gradient(135deg, var(--primary) 0%, #2d5a87 100%);
    color: white;
    padding: 1.5rem 2rem;
    position: sticky;
    top: 0;
    z-index: 100;
    box-shadow: 0 2px 10px rgba(0,0,0,0.15);
}

header h1 {
    font-size: 1.75rem;
    font-weight: 600;
    margin-bottom: 0.5rem;
}

.doc-info {
    display: flex;
    gap: 2rem;
    font-size: 0.9rem;
    opacity: 0.9;
    flex-wrap: wrap;
}

.doc-info span {
    display: flex;
    align-items: center;
    gap: 0.5rem;
}

.search-container {
    margin-top: 1rem;
    display: flex;
    align-items: center;
    gap: 1rem;
}

#searchBox {
    width: 100%;
    max-width: 400px;
    padding: 0.75rem 1rem;
    font-size: 1rem;
    border: none;
    border-radius: 6px;
    outline: none;
    transition: box-shadow 0.2s;
}

#searchBox:focus {
    box-shadow: 0 0 0 3px rgba(255,255,255,0.3);
}

#searchResults {
    font-size: 0.9rem;
    opacity: 0.9;
}

/* Sidebar Navigation */
nav#sidebar {
    position: fixed;
    left: 0;
    top: 140px;
    width: 240px;
    height: calc(100vh - 140px);
    background: var(--white);
    border-right: 1px solid var(--border);
    padding: 1.5rem 1rem;
    overflow-y: auto;
    z-index: 50;
}

nav#sidebar h3 {
    font-size: 0.85rem;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    color: var(--text-light);
    margin-bottom: 1rem;
    padding-bottom: 0.5rem;
    border-bottom: 2px solid var(--primary);
}

nav#sidebar ul {
    list-style: none;
}

nav#sidebar li {
    margin-bottom: 0.25rem;
}

nav#sidebar a {
    display: block;
    padding: 0.6rem 0.75rem;
    color: var(--text);
    text-decoration: none;
    border-radius: 4px;
    font-size: 0.9rem;
    transition: all 0.2s;
}

nav#sidebar a:hover {
    background: var(--accent);
    color: var(--primary);
}

nav#sidebar .count {
    float: right;
    background: var(--accent);
    color: var(--primary);
    padding: 0.1rem 0.5rem;
    border-radius: 10px;
    font-size: 0.8rem;
    font-weight: 600;
}

/* Main Content */
main {
    margin-left: 260px;
    padding: 2rem;
    max-width: 1400px;
}

/* Category Sections */
.category-section {
    background: var(--white);
    border-radius: 8px;
    margin-bottom: 1.5rem;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    overflow: hidden;
}

.category-header {
    background: var(--accent);
    padding: 1rem 1.5rem;
    cursor: pointer;
    display: flex;
    justify-content: space-between;
    align-items: center;
    user-select: none;
    transition: background 0.2s;
}

.category-header:hover {
    background: var(--accent-dark);
}

.category-header h2 {
    font-size: 1.1rem;
    font-weight: 600;
    color: var(--primary);
    display: flex;
    align-items: center;
    gap: 0.75rem;
}

.category-header .toggle-icon {
    font-size: 0.8rem;
    transition: transform 0.2s;
}

.category-header .table-count {
    font-size: 0.85rem;
    color: var(--text-light);
}

.category-content {
    padding: 1.5rem;
}

.category-content.collapsed {
    display: none;
}

/* Table Blocks */
.table-block {
    margin-bottom: 1.5rem;
    border: 1px solid var(--border);
    border-radius: 6px;
    overflow: hidden;
}

.table-block:last-child {
    margin-bottom: 0;
}

.table-header {
    background: linear-gradient(to right, #f8f9fa, #fff);
    padding: 0.75rem 1rem;
    display: flex;
    justify-content: space-between;
    align-items: center;
    border-bottom: 1px solid var(--border);
}

.table-title {
    font-weight: 600;
    color: var(--text);
    font-size: 0.95rem;
}

.table-id {
    font-size: 0.8rem;
    color: var(--text-light);
    margin-left: 0.5rem;
}

.page-ref {
    background: var(--primary);
    color: white;
    padding: 0.25rem 0.75rem;
    border-radius: 4px;
    font-size: 0.85rem;
    font-weight: 500;
}

.page-ref a {
    color: white;
    text-decoration: none;
}

/* Data Tables */
.table-container {
    overflow-x: auto;
    max-height: 500px;
    overflow-y: auto;
}

table {
    width: 100%;
    border-collapse: collapse;
    font-size: 0.9rem;
}

th, td {
    border: 1px solid var(--border);
    padding: 0.6rem 0.75rem;
    text-align: left;
    vertical-align: top;
}

th {
    background: var(--primary);
    color: white;
    font-weight: 600;
    position: sticky;
    top: 0;
    z-index: 10;
}

tr:nth-child(even) {
    background: #f8f9fa;
}

tr:hover {
    background: var(--accent);
}

td {
    max-width: 400px;
    word-wrap: break-word;
}

/* Search Highlighting */
.highlight {
    background: #ffeb3b;
    padding: 0.1rem 0.2rem;
    border-radius: 2px;
}

/* Empty State */
.empty-state {
    text-align: center;
    padding: 3rem;
    color: var(--text-light);
}

.empty-state h3 {
    margin-bottom: 0.5rem;
}

/* Footer */
footer {
    margin-left: 260px;
    padding: 2rem;
    text-align: center;
    color: var(--text-light);
    font-size: 0.85rem;
    border-top: 1px solid var(--border);
    background: var(--white);
}

/* Print Styles */
@media print {
    nav#sidebar {
        display: none;
    }

    main, footer {
        margin-left: 0;
    }

    header {
        position: relative;
        background: var(--primary);
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
    }

    .category-section {
        break-inside: avoid;
        page-break-inside: avoid;
    }

    .table-block {
        break-inside: avoid;
        page-break-inside: avoid;
    }

    .category-content.collapsed {
        display: block !important;
    }
}

/* Collapsible Table Rows */
.table-row-group {
    display: contents;
}

.table-row-header {
    cursor: pointer;
    background: var(--accent-dark) !important;
    font-weight: 600;
    user-select: none;
}

.table-row-header:hover {
    background: var(--secondary) !important;
    color: white;
}

.table-row-header td {
    padding: 0.6rem 0.75rem !important;
}

.table-row-header .toggle-row-icon {
    display: inline-block;
    margin-right: 0.5rem;
    transition: transform 0.2s ease;
    font-size: 0.85rem;
}

.table-row-header.collapsed .toggle-row-icon {
    transform: rotate(-90deg);
}

.table-data-row {
    display: table-row;
}

.table-data-row.hidden {
    display: none;
}

/* Responsive */
@media (max-width: 900px) {
    nav#sidebar {
        position: relative;
        width: 100%;
        height: auto;
        top: 0;
        border-right: none;
        border-bottom: 1px solid var(--border);
    }

    main, footer {
        margin-left: 0;
    }
}
"""

# =============================================================================
# JavaScript (Inline)
# =============================================================================
JS_SCRIPTS = """
// Toggle category section collapse/expand
function toggleSection(id) {
    const content = document.getElementById('content-' + id);
    const icon = document.getElementById('icon-' + id);

    if (content.classList.contains('collapsed')) {
        content.classList.remove('collapsed');
        icon.textContent = '▼';
        icon.style.transform = 'rotate(0deg)';
    } else {
        content.classList.add('collapsed');
        icon.textContent = '▶';
        icon.style.transform = 'rotate(0deg)';
    }
}

// Expand all sections
function expandAll() {
    document.querySelectorAll('.category-content').forEach(el => {
        el.classList.remove('collapsed');
    });
    document.querySelectorAll('.toggle-icon').forEach(el => {
        el.textContent = '▼';
    });
}

// Collapse all sections
function collapseAll() {
    document.querySelectorAll('.category-content').forEach(el => {
        el.classList.add('collapsed');
    });
    document.querySelectorAll('.toggle-icon').forEach(el => {
        el.textContent = '▶';
    });
}

// Global search functionality
function searchTables() {
    const query = document.getElementById('searchBox').value.toLowerCase().trim();
    const tableBlocks = document.querySelectorAll('.table-block');
    const categorySections = document.querySelectorAll('.category-section');
    let matchCount = 0;
    let tableMatchCount = 0;

    // Remove previous highlights
    document.querySelectorAll('.highlight').forEach(el => {
        const parent = el.parentNode;
        parent.replaceChild(document.createTextNode(el.textContent), el);
        parent.normalize();
    });

    // Reset visibility
    if (query.length < 2) {
        document.getElementById('searchResults').textContent = '';
        tableBlocks.forEach(t => t.style.display = 'block');
        categorySections.forEach(s => s.style.display = 'block');
        return;
    }

    // Search and filter
    categorySections.forEach(section => {
        const blocks = section.querySelectorAll('.table-block');
        let sectionHasMatch = false;

        blocks.forEach(block => {
            const text = block.textContent.toLowerCase();
            if (text.includes(query)) {
                block.style.display = 'block';
                sectionHasMatch = true;
                tableMatchCount++;

                // Highlight matches in cells
                const cells = block.querySelectorAll('td, th');
                cells.forEach(cell => {
                    if (cell.textContent.toLowerCase().includes(query)) {
                        highlightText(cell, query);
                        matchCount++;
                    }
                });
            } else {
                block.style.display = 'none';
            }
        });

        // Show/hide section based on matches
        if (sectionHasMatch) {
            section.style.display = 'block';
            // Expand section with matches
            const content = section.querySelector('.category-content');
            const icon = section.querySelector('.toggle-icon');
            if (content) content.classList.remove('collapsed');
            if (icon) icon.textContent = '▼';
        } else {
            section.style.display = 'none';
        }
    });

    document.getElementById('searchResults').textContent =
        `Found ${tableMatchCount} table(s) with ${matchCount} matching cell(s)`;
}

// Toggle table row group (collapsible rows)
function toggleTableRowGroup(headerId) {
    const headerRow = document.getElementById(headerId);
    if (!headerRow) return;

    const table = headerRow.closest('table');
    if (!table) return;

    const rowIndex = Array.from(headerRow.parentNode.children).indexOf(headerRow);
    const icon = headerRow.querySelector('.toggle-row-icon');
    const isCollapsed = headerRow.classList.contains('collapsed');

    // Find all subsequent data rows for this group
    let currentRow = headerRow.nextElementSibling;
    let rowsToToggle = [];

    while (currentRow && currentRow.classList.contains('table-data-row')) {
        rowsToToggle.push(currentRow);
        currentRow = currentRow.nextElementSibling;
    }

    // Toggle visibility
    if (isCollapsed) {
        headerRow.classList.remove('collapsed');
        rowsToToggle.forEach(row => row.classList.remove('hidden'));
        if (icon) icon.textContent = '▼';
    } else {
        headerRow.classList.add('collapsed');
        rowsToToggle.forEach(row => row.classList.add('hidden'));
        if (icon) icon.textContent = '▶';
    }
}

// Highlight text in element
function highlightText(element, query) {
    const innerHTML = element.innerHTML;
    const regex = new RegExp(`(${escapeRegex(query)})`, 'gi');
    element.innerHTML = innerHTML.replace(regex, '<span class="highlight">$1</span>');
}

// Escape regex special characters
function escapeRegex(string) {
    return string.replace(/[.*+?^${}()|[\\]\\\\]/g, '\\\\$&');
}

// Smooth scroll to category
function scrollToCategory(id) {
    const element = document.getElementById(id);
    if (element) {
        element.scrollIntoView({ behavior: 'smooth', block: 'start' });
        // Expand the section
        const content = element.querySelector('.category-content');
        const icon = element.querySelector('.toggle-icon');
        if (content) content.classList.remove('collapsed');
        if (icon) icon.textContent = '▼';
    }
}

// Initialize on load
document.addEventListener('DOMContentLoaded', function() {
    // Add click handlers for sidebar links
    document.querySelectorAll('nav#sidebar a').forEach(link => {
        link.addEventListener('click', function(e) {
            e.preventDefault();
            const targetId = this.getAttribute('href').substring(1);
            scrollToCategory(targetId);
        });
    });
});
"""


# =============================================================================
# Core Functions
# =============================================================================

def load_meta(input_path: Path) -> Dict:
    """
    Load meta.json and extract table data with validation.

    Args:
        input_path: Path to the _meta.json file

    Returns:
        Dictionary containing document metadata and tables

    Raises:
        FileNotFoundError: If input file doesn't exist
        ValueError: If required fields are missing
    """
    if not input_path.exists():
        raise FileNotFoundError(f"Meta file not found: {input_path}")

    with open(input_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Validate required fields
    if 'tables' not in data:
        raise ValueError("meta.json missing 'tables' field")

    # Ensure tables is a list
    if not isinstance(data['tables'], list):
        raise ValueError("'tables' field must be a list")

    print(f"  Loaded {len(data['tables'])} tables from meta.json")

    return data


def parse_markdown_table(content: str) -> Tuple[List[str], List[List[str]]]:
    """
    Parse markdown pipe-delimited table into headers and rows.

    Args:
        content: Markdown table string with pipe delimiters

    Returns:
        Tuple of (headers, rows) where headers is a list of column names
        and rows is a list of row data (each row is a list of cell values)
    """
    if not content or not content.strip():
        return [], []

    lines = [l.strip() for l in content.strip().split('\n') if l.strip()]

    headers = []
    rows = []
    header_found = False

    for i, line in enumerate(lines):
        # Skip lines that don't look like table rows
        if not line.startswith('|'):
            continue

        # Skip separator lines (|---|---|)
        clean_line = line.replace('|', '').strip()
        if set(clean_line) <= {'-', ':', ' '}:
            continue

        # Parse cells
        cells = [c.strip() for c in line.split('|')]
        # Remove empty first/last elements from split
        if cells and cells[0] == '':
            cells = cells[1:]
        if cells and cells[-1] == '':
            cells = cells[:-1]

        if not header_found:
            headers = cells
            header_found = True
        else:
            rows.append(cells)

    return headers, rows


def categorize_tables(tables: List[Dict]) -> Dict[str, List[Dict]]:
    """
    Group tables by inferred ESIA category based on content analysis.

    Categories follow standard ESIA structure:
    - Contents & Overview
    - Legal & Regulatory Framework
    - Physical Environment (Baseline)
    - Biological Environment (Baseline)
    - Social Environment (Baseline)
    - Impact Assessment
    - Mitigation & Management
    - Monitoring
    - Closure & Rehabilitation
    - Other Tables

    Args:
        tables: List of table dictionaries from meta.json

    Returns:
        OrderedDict mapping category names to lists of tables
    """
    # Define categories with keywords (order matters - first match wins)
    category_patterns = [
        ('Contents & Overview', [
            'contents', 'table of contents', 'list of tables', 'list of figures',
            'abbreviation', 'acronym', 'glossary', 'executive summary', 'overview'
        ]),
        ('Legal & Regulatory', [
            'legislation', 'regulation', 'permit', 'compliance', 'legal', 'law',
            'standard', 'guideline', 'requirement', 'policy', 'act', 'decree',
            'ifc', 'world bank', 'adb', 'safeguard'
        ]),
        ('Physical Environment', [
            'climate', 'meteorolog', 'hydrology', 'hydrogeolog', 'geology', 'geolog',
            'soil', 'topograph', 'air quality', 'water quality', 'noise', 'vibration',
            'seismic', 'erosion', 'sediment', 'groundwater', 'surface water',
            'rainfall', 'temperature', 'wind', 'humidity'
        ]),
        ('Biological Environment', [
            'flora', 'fauna', 'species', 'habitat', 'biodiversity', 'ecology',
            'wildlife', 'vegetation', 'forest', 'tree', 'bird', 'mammal', 'reptile',
            'amphibian', 'fish', 'invertebrate', 'endangered', 'protected', 'iucn',
            'critical habitat', 'ecosystem'
        ]),
        ('Social Environment', [
            'population', 'community', 'livelihood', 'employment', 'health',
            'education', 'demographic', 'household', 'income', 'gender',
            'indigenous', 'resettlement', 'displacement', 'land use', 'land tenure',
            'stakeholder', 'consultation', 'grievance', 'cultural heritage',
            'archaeological', 'ethnographic'
        ]),
        ('Impact Assessment', [
            'impact', 'significance', 'rating', 'assessment', 'evaluation',
            'magnitude', 'severity', 'probability', 'likelihood', 'consequence',
            'cumulative', 'transboundary', 'residual'
        ]),
        ('Mitigation & Management', [
            'mitigation', 'management', 'measure', 'commitment', 'action',
            'esmp', 'emp', 'esia management', 'environmental management',
            'social management', 'control', 'prevention', 'minimize'
        ]),
        ('Monitoring', [
            'monitoring', 'indicator', 'parameter', 'frequency', 'sampling',
            'measurement', 'surveillance', 'audit', 'inspection', 'reporting',
            'kpi', 'performance'
        ]),
        ('Closure & Rehabilitation', [
            'closure', 'rehabilitation', 'decommission', 'restoration',
            'reclamation', 'post-closure', 'abandonment', 'remediation'
        ])
    ]

    # Initialize categories (preserving order)
    categories = {cat: [] for cat, _ in category_patterns}
    categories['Other Tables'] = []

    for table in tables:
        content_lower = table.get('content', '').lower()
        caption_lower = (table.get('caption') or '').lower()
        search_text = content_lower + ' ' + caption_lower

        matched = False
        for category, keywords in category_patterns:
            if any(kw in search_text for kw in keywords):
                categories[category].append(table)
                matched = True
                break

        if not matched:
            categories['Other Tables'].append(table)

    # Remove empty categories (preserving order)
    return {k: v for k, v in categories.items() if v}


def sanitize_sheet_name(name: str) -> str:
    """
    Sanitize string for use as Excel sheet name.

    Excel sheet names have restrictions:
    - Max 31 characters
    - Cannot contain: \\ / ? * [ ]
    - Cannot be blank

    Args:
        name: Raw category/sheet name

    Returns:
        Sanitized sheet name safe for Excel
    """
    # Remove/replace invalid characters
    sanitized = re.sub(r'[\\/?*\[\]]', '', name)
    sanitized = sanitized.replace('&', 'and')
    # Truncate to 31 chars
    sanitized = sanitized[:31].strip()
    # Ensure not empty
    return sanitized or 'Sheet'


def escape_html_content(text: str) -> str:
    """Escape HTML special characters in text content."""
    if text is None:
        return ''
    return html.escape(str(text))


def generate_table_html(table: Dict, table_index: int) -> str:
    """
    Generate HTML for a single table block.

    Args:
        table: Table dictionary from meta.json
        table_index: Index for unique ID generation

    Returns:
        HTML string for the table block
    """
    table_id = table.get('table_id', table_index)
    page = table.get('page', 'N/A')
    position = table.get('position', f'table_{table_id}')
    caption = table.get('caption', '')
    content = table.get('content', '')

    # Parse markdown table
    headers, rows = parse_markdown_table(content)

    # Generate title
    if caption:
        title = escape_html_content(caption)
    else:
        # Try to infer title from first row content
        if headers:
            title = f"Table {table_id}"
        else:
            title = f"Table {table_id}"

    # Build HTML
    html_parts = [
        f'<div class="table-block" data-table-id="{table_id}" data-page="{page}" id="table-{position}">',
        f'  <div class="table-header">',
        f'    <span class="table-title">{title}<span class="table-id">(ID: {position})</span></span>',
        f'    <span class="page-ref">Page {page}</span>',
        f'  </div>',
        f'  <div class="table-container">'
    ]

    if headers or rows:
        html_parts.append('    <table>')

        # Headers
        if headers:
            html_parts.append('      <thead><tr>')
            for header in headers:
                html_parts.append(f'        <th>{escape_html_content(header)}</th>')
            html_parts.append('      </tr></thead>')

        # Body rows with collapsible groups
        if rows:
            html_parts.append('      <tbody>')

            # Group rows into collapsible sections (5 rows per group)
            rows_per_group = 5
            for group_idx, i in enumerate(range(0, len(rows), rows_per_group)):
                group_rows = rows[i:i + rows_per_group]
                group_id = f"table-{position}-group-{group_idx}"

                # Add group header (collapsible)
                first_row = group_rows[0]
                preview_text = ' | '.join(first_row[:2])[:50]  # Show first 2 columns
                if len(preview_text) > 45:
                    preview_text = preview_text[:45] + '...'

                html_parts.append(f'        <tr id="{group_id}" class="table-row-header" onclick="toggleTableRowGroup(\'{group_id}\')">')
                html_parts.append(f'          <td colspan="{len(headers) if headers else len(first_row)}" style="cursor: pointer; font-weight: 600;">')
                html_parts.append(f'            <span class="toggle-row-icon">▼</span>Rows {i+1}-{min(i+rows_per_group, len(rows))} <span style="color: var(--text-light); font-size: 0.85em;">({preview_text})</span>')
                html_parts.append('          </td>')
                html_parts.append('        </tr>')

                # Add data rows for this group
                for row_offset, row in enumerate(group_rows):
                    html_parts.append(f'        <tr class="table-data-row">')
                    for cell in row:
                        html_parts.append(f'          <td>{escape_html_content(cell)}</td>')
                    # Pad with empty cells if row is shorter than headers
                    if headers and len(row) < len(headers):
                        for _ in range(len(headers) - len(row)):
                            html_parts.append('          <td></td>')
                    html_parts.append('        </tr>')

            html_parts.append('      </tbody>')

        html_parts.append('    </table>')
    else:
        # Raw content fallback
        html_parts.append(f'    <pre style="padding: 1rem; background: #f8f9fa; overflow-x: auto;">{escape_html_content(content)}</pre>')

    html_parts.extend([
        '  </div>',
        '</div>'
    ])

    return '\n'.join(html_parts)


def generate_category_sections(categorized: Dict[str, List[Dict]]) -> str:
    """
    Generate HTML for all category sections with their tables.

    Args:
        categorized: Dictionary mapping categories to table lists

    Returns:
        HTML string for all category sections
    """
    sections_html = []

    for cat_index, (category, tables) in enumerate(categorized.items()):
        cat_id = f"cat-{cat_index}"

        section_html = [
            f'<section class="category-section" id="{cat_id}">',
            f'  <div class="category-header" onclick="toggleSection({cat_index})">',
            f'    <h2><span class="toggle-icon" id="icon-{cat_index}">▼</span>{escape_html_content(category)}</h2>',
            f'    <span class="table-count">{len(tables)} table(s)</span>',
            f'  </div>',
            f'  <div class="category-content" id="content-{cat_index}">'
        ]

        # Add each table
        for i, table in enumerate(tables):
            section_html.append(generate_table_html(table, i))

        section_html.extend([
            '  </div>',
            '</section>'
        ])

        sections_html.append('\n'.join(section_html))

    return '\n\n'.join(sections_html)


def build_html(data: Dict, output_path: Path):
    """
    Generate the interactive HTML fact browser.

    Args:
        data: Full meta.json data dictionary
        output_path: Path for output HTML file
    """
    tables = data.get('tables', [])
    doc_info = data.get('document', {})
    stats = data.get('statistics', {})

    # Categorize tables
    categorized = categorize_tables(tables)

    # Document info
    filename = doc_info.get('original_filename', 'Unknown Document')
    total_pages = doc_info.get('total_pages', 'N/A')
    total_tables = stats.get('total_tables', len(tables))

    # Build sidebar navigation
    sidebar_items = []
    for i, (cat, tbls) in enumerate(categorized.items()):
        sidebar_items.append(
            f'<li><a href="#cat-{i}">{escape_html_content(cat)}<span class="count">{len(tbls)}</span></a></li>'
        )
    sidebar_html = '\n            '.join(sidebar_items)

    # Generate category sections
    content_html = generate_category_sections(categorized)

    # Build full HTML
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ESIA Fact Browser - {escape_html_content(filename)}</title>
    <style>
{CSS_STYLES}
    </style>
</head>
<body>
    <header>
        <h1>ESIA Table Fact Browser</h1>
        <div class="doc-info">
            <span>📄 {escape_html_content(filename)}</span>
            <span>📑 {total_pages} pages</span>
            <span>📊 {total_tables} tables</span>
        </div>
        <div class="search-container">
            <input type="text" id="searchBox" placeholder="Search all tables..." onkeyup="searchTables()">
            <span id="searchResults"></span>
        </div>
    </header>

    <nav id="sidebar">
        <h3>Categories</h3>
        <ul>
            {sidebar_html}
        </ul>
        <div style="margin-top: 1rem; padding-top: 1rem; border-top: 1px solid var(--border);">
            <button onclick="expandAll()" style="width: 100%; margin-bottom: 0.5rem; padding: 0.5rem; cursor: pointer;">Expand All</button>
            <button onclick="collapseAll()" style="width: 100%; padding: 0.5rem; cursor: pointer;">Collapse All</button>
        </div>
    </nav>

    <main>
        {content_html}
    </main>

    <footer>
        <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')} | ESIA Table Fact Browser</p>
        <p>Tables extracted from: {escape_html_content(filename)}</p>
    </footer>

    <script>
{JS_SCRIPTS}
    </script>
</body>
</html>'''

    # Write to file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"  Generated HTML: {output_path}")
    print(f"  - {len(categorized)} categories")
    print(f"  - {total_tables} tables")


def build_excel(data: Dict, output_path: Path):
    """
    Generate Excel workbook with all tables organized by category.

    Args:
        data: Full meta.json data dictionary
        output_path: Path for output Excel file
    """
    if not EXCEL_AVAILABLE:
        print("  Skipping Excel export - openpyxl not installed")
        return

    tables = data.get('tables', [])
    doc_info = data.get('document', {})
    stats = data.get('statistics', {})

    # Categorize tables
    categorized = categorize_tables(tables)

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Define styles
    header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    table_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    table_header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ----- Summary Sheet -----
    summary_ws = wb.create_sheet("Summary")

    summary_ws.cell(1, 1, "ESIA Table Fact Browser - Summary")
    summary_ws.cell(1, 1).font = title_font
    summary_ws.merge_cells('A1:D1')

    summary_data = [
        ("Document", doc_info.get('original_filename', 'Unknown')),
        ("Total Pages", doc_info.get('total_pages', 'N/A')),
        ("Total Tables", stats.get('total_tables', len(tables))),
        ("Generated", datetime.now().strftime('%Y-%m-%d %H:%M')),
        ("", ""),
        ("Category", "Table Count")
    ]

    row = 3
    for label, value in summary_data:
        summary_ws.cell(row, 1, label)
        summary_ws.cell(row, 2, value)
        if label:
            summary_ws.cell(row, 1).font = Font(bold=True)
        row += 1

    # Category counts
    for category, cat_tables in categorized.items():
        summary_ws.cell(row, 1, category)
        summary_ws.cell(row, 2, len(cat_tables))
        row += 1

    # Adjust column widths
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 50

    # ----- Category Sheets -----
    for category, cat_tables in categorized.items():
        sheet_name = sanitize_sheet_name(category)
        ws = wb.create_sheet(sheet_name)

        current_row = 1

        for table in cat_tables:
            table_id = table.get('table_id', 0)
            page = table.get('page', 'N/A')
            position = table.get('position', f'table_{table_id}')
            caption = table.get('caption', '')
            content = table.get('content', '')

            # Table header row
            title = caption if caption else f"Table {table_id}"
            ws.cell(current_row, 1, f"{title} (Page {page})")
            ws.cell(current_row, 1).font = subtitle_font
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            current_row += 1

            # Parse and write table content
            headers, rows = parse_markdown_table(content)

            if headers:
                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(current_row, col_idx, header)
                    cell.fill = table_header_fill
                    cell.font = table_header_font
                    cell.border = thin_border

                # Add Source_Page column header
                page_col = len(headers) + 1
                cell = ws.cell(current_row, page_col, "Source_Page")
                cell.fill = table_header_fill
                cell.font = table_header_font
                cell.border = thin_border
                current_row += 1

                # Write data rows
                for row_data in rows:
                    for col_idx, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(current_row, col_idx, cell_value)
                        cell.border = thin_border
                    # Add page number in last column
                    ws.cell(current_row, page_col, page).border = thin_border
                    current_row += 1
            else:
                # Raw content fallback
                ws.cell(current_row, 1, content[:1000])  # Truncate very long content
                ws.cell(current_row, 2, page)
                current_row += 1

            # Spacing between tables
            current_row += 2

        # Auto-adjust column widths (approximate)
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 20

    # Freeze header rows
    for ws in wb.worksheets:
        if ws.title != "Summary":
            ws.freeze_panes = 'A2'

    # Save workbook
    wb.save(output_path)
    print(f"  Generated Excel: {output_path}")
    print(f"  - {len(categorized) + 1} worksheets")


def main():
    """Main CLI entry point."""
    parser = argparse.ArgumentParser(
        description='Generate ESIA Table Fact Browser from meta.json',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python build_fact_browser.py --input document_meta.json --output ./output/
    python build_fact_browser.py -i data/outputs/ESIA_Report_meta.json -o data/outputs/
        """
    )
    parser.add_argument(
        '--input', '-i',
        required=True,
        help='Path to _meta.json file'
    )
    parser.add_argument(
        '--output', '-o',
        required=True,
        help='Output directory for generated files'
    )
    parser.add_argument(
        '--html-only',
        action='store_true',
        help='Generate only HTML output (skip Excel)'
    )
    parser.add_argument(
        '--excel-only',
        action='store_true',
        help='Generate only Excel output (skip HTML)'
    )

    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output)

    print("=" * 60)
    print("ESIA TABLE FACT BROWSER GENERATOR")
    print("=" * 60)
    print()

    # Validate input
    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}")
        return 1

    # Create output directory
    output_dir.mkdir(parents=True, exist_ok=True)

    # Derive stem from input filename
    stem = input_path.stem.replace('_meta', '')

    # Load data
    print(f"Loading: {input_path}")
    try:
        data = load_meta(input_path)
    except Exception as e:
        print(f"Error loading meta.json: {e}")
        return 1

    tables = data.get('tables', [])
    print(f"  Found {len(tables)} tables to process")
    print()

    # Generate outputs
    if not args.excel_only:
        html_path = output_dir / f"{stem}_fact_browser.html"
        print(f"Generating HTML...")
        try:
            build_html(data, html_path)
        except Exception as e:
            print(f"Error generating HTML: {e}")
            return 1
        print()

    if not args.html_only:
        xlsx_path = output_dir / f"{stem}_fact_browser.xlsx"
        print(f"Generating Excel...")
        try:
            build_excel(data, xlsx_path)
        except Exception as e:
            print(f"Error generating Excel: {e}")
            return 1
        print()

    print("=" * 60)
    print("GENERATION COMPLETE")
    print("=" * 60)
    print(f"Output directory: {output_dir}")

    return 0


if __name__ == '__main__':
    exit(main())
