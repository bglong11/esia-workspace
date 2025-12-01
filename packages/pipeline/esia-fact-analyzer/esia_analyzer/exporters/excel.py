"""
Excel workbook export for ESIA analysis results.
"""

from pathlib import Path
from typing import Dict, List, Any

# Optional imports with graceful fallback
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def export_excel(
    output_path: Path,
    summary: Dict,
    categorized_facts: Dict,
    issues: List[Dict],
    unit_issues: List[Dict],
    threshold_checks: List[Dict],
    gaps: List[Dict],
    factsheet_summary: Dict = None
):
    """
    Export analysis results to Excel workbook.

    Args:
        output_path: Path to save Excel file
        summary: Analysis summary dictionary
        categorized_facts: Facts organized by category
        issues: List of consistency issues
        unit_issues: List of unit standardization issues
        threshold_checks: List of threshold check results
        gaps: List of gap analysis results
        factsheet_summary: Optional LLM-generated factsheet summary
    """
    if not HAS_OPENPYXL:
        print("Warning: openpyxl not installed. Skipping Excel export.")
        return

    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    red_text = Font(color="721c24")
    orange_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    orange_text = Font(color="856404")
    green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    green_text = Font(color="155724")
    blue_fill = PatternFill(start_color="cce5ff", end_color="cce5ff", fill_type="solid")

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    row = 1
    ws_summary.cell(row, 1, "ESIA Review Analysis Summary").font = Font(size=16, bold=True)
    row += 2

    for key, value in summary.items():
        if isinstance(value, dict):
            ws_summary.cell(row, 1, key.replace("_", " ").title()).font = Font(bold=True)
            row += 1
            for k, v in value.items():
                ws_summary.cell(row, 1, f"  {k}")
                ws_summary.cell(row, 2, str(v))
                row += 1
        else:
            ws_summary.cell(row, 1, key.replace("_", " ").title())
            ws_summary.cell(row, 2, str(value))
            row += 1

    # Project Summary sheet (if available)
    if factsheet_summary and factsheet_summary.get('paragraphs'):
        ws_factsheet = wb.create_sheet("Project Summary")
        ws_factsheet.append(["Section", "Content"])
        for cell in ws_factsheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        section_titles = {
            'project_overview': 'Project Overview',
            'baseline': 'Environmental & Social Baseline',
            'impacts': 'Major Anticipated Impacts',
            'mitigation': 'Mitigation & Management Measures',
            'residual_risks': 'Residual Risks & Compliance'
        }

        for key, title in section_titles.items():
            content = factsheet_summary.get('paragraphs', {}).get(key, '')
            if content:
                ws_factsheet.append([title, content])

    # Categories sheet
    ws_cat = wb.create_sheet("Fact Categories")
    ws_cat.append(["Category", "Count", "Sample Sections"])
    for cell in ws_cat[1]:
        cell.fill = header_fill
        cell.font = header_font

    for cat, facts in sorted(categorized_facts.items(), key=lambda x: -len(x[1])):
        sections = list(set([(f.get("section") or "Unknown")[:50] for f in facts[:5]]))
        ws_cat.append([cat, len(facts), "; ".join(sections)])

    # Consistency Issues sheet
    ws_issues = wb.create_sheet("Consistency Issues")
    ws_issues.append(["Severity", "Parameter Context", "Values Found", "Normalized (base unit)", "Difference %", "Details"])
    for cell in ws_issues[1]:
        cell.fill = header_fill
        cell.font = header_font

    for issue in issues:
        row_data = [
            issue.get("severity", "").upper(),
            issue.get("context", ""),
            ", ".join(issue.get("values", [])),
            f"{issue.get('normalized_values', [])} ({issue.get('base_unit', '')})",
            f"{issue.get('diff_percent', 0)}%",
            issue.get("message", "")
        ]
        ws_issues.append(row_data)
        row_idx = ws_issues.max_row
        if issue.get("severity") == "high":
            for cell in ws_issues[row_idx]:
                cell.fill = red_fill
                cell.font = red_text
        elif issue.get("severity") == "medium":
            for cell in ws_issues[row_idx]:
                cell.fill = orange_fill
                cell.font = orange_text

    # Unit Standardization sheet
    ws_units = wb.create_sheet("Unit Standardization")
    ws_units.append(["Parameter Context", "Units Used", "Examples", "Recommendation"])
    for cell in ws_units[1]:
        cell.fill = header_fill
        cell.font = header_font

    for issue in unit_issues:
        examples_str = "; ".join([f"{e['value']} {e['unit']} (p.{e['page']})" for e in issue.get('examples', [])])
        recommended_unit = issue.get('base_unit', '')
        row_data = [
            issue.get("context", ""),
            ", ".join(issue.get("units_used", [])),
            examples_str,
            f"Standardize to {recommended_unit}"
        ]
        ws_units.append(row_data)
        row_idx = ws_units.max_row
        for cell in ws_units[row_idx]:
            cell.fill = orange_fill
            cell.font = orange_text

    # Threshold Checks sheet
    ws_thresh = wb.create_sheet("Threshold Compliance")
    ws_thresh.append(["Parameter", "Category", "Value", "Threshold", "Unit", "Status", "Page", "Source"])
    for cell in ws_thresh[1]:
        cell.fill = header_fill
        cell.font = header_font

    for check in threshold_checks:
        ws_thresh.append([
            check.get("parameter"),
            check.get("category"),
            check.get("value"),
            check.get("threshold"),
            check.get("unit"),
            check.get("status"),
            check.get("page"),
            check.get("source")
        ])
        row_idx = ws_thresh.max_row
        status = check.get("status")
        if status == "EXCEEDANCE":
            ws_thresh.cell(row_idx, 6).fill = red_fill
            ws_thresh.cell(row_idx, 6).font = red_text
        elif status == "APPROACHING":
            ws_thresh.cell(row_idx, 6).fill = orange_fill
            ws_thresh.cell(row_idx, 6).font = orange_text
        elif status == "COMPLIANT":
            ws_thresh.cell(row_idx, 6).fill = green_fill
            ws_thresh.cell(row_idx, 6).font = green_text

    # Gap Analysis sheet
    ws_gaps = wb.create_sheet("Gap Analysis")
    ws_gaps.append(["Section", "Sub-section", "Status", "Content Found", "Page(s)"])
    for cell in ws_gaps[1]:
        cell.fill = header_fill
        cell.font = header_font

    for gap in gaps:
        matches = gap.get("matches", [])
        if matches:
            content_text = "; ".join([m.get("content", "")[:100] for m in matches])
            pages_text = ", ".join([str(m.get("page", "?")) for m in matches])
        else:
            content_text = ""
            pages_text = ""

        ws_gaps.append([
            gap.get("section"),
            gap.get("item"),
            gap.get("status"),
            content_text,
            pages_text
        ])
        row_idx = ws_gaps.max_row
        if gap.get("status") == "MISSING":
            ws_gaps.cell(row_idx, 3).fill = red_fill
            ws_gaps.cell(row_idx, 3).font = red_text
        else:
            ws_gaps.cell(row_idx, 3).fill = green_fill
            ws_gaps.cell(row_idx, 3).font = green_text

    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")
