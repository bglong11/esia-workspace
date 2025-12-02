#!/usr/bin/env python3
"""
ESIA Review Factsheet Generator

Generates an ESIA review factsheet in both Excel and HTML format from pipeline outputs.

Input files:
    - *_facts.json: Structured facts extracted from the ESIA
    - *_meta.json: Document metadata and extracted tables
    - *_chunks.jsonl: Chunked text of the ESIA

Output files:
    - <base_name>_ESIA_review.xlsx: Excel workbook with analysis sheets
    - <base_name>_ESIA_review.html: Interactive HTML dashboard

Usage:
    python generate_esia_factsheet.py

    Modify the file paths at the top of the script to match your input files.
"""

import json
import re
import os
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Dict, List, Any, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION - Modify these paths to match your input files
# =============================================================================

FACTS_PATH = Path("data/outputs/1764662314649-DUMMY Lake Toba ESIA_facts.json")
META_PATH = Path("data/outputs/1764662314649-DUMMY Lake Toba ESIA_meta.json")
CHUNKS_PATH = Path("data/outputs/1764662314649-DUMMY Lake Toba ESIA_chunks.jsonl")

# =============================================================================
# CONSTANTS
# =============================================================================

# Unit conversion factors to base units
UNIT_CONVERSIONS = {
    # Area units -> sq m
    'ha': {'base': 'sq m', 'factor': 10000},
    'hectare': {'base': 'sq m', 'factor': 10000},
    'hectares': {'base': 'sq m', 'factor': 10000},
    'km²': {'base': 'sq m', 'factor': 1000000},
    'km2': {'base': 'sq m', 'factor': 1000000},
    'square kilometers': {'base': 'sq m', 'factor': 1000000},
    'm²': {'base': 'sq m', 'factor': 1},
    'm2': {'base': 'sq m', 'factor': 1},
    'sq m': {'base': 'sq m', 'factor': 1},
    'sqm': {'base': 'sq m', 'factor': 1},
    'acres': {'base': 'sq m', 'factor': 4046.86},
    'acre': {'base': 'sq m', 'factor': 4046.86},
    'sq km': {'base': 'sq m', 'factor': 1000000},
    'sq ft': {'base': 'sq m', 'factor': 0.092903},
    'ft²': {'base': 'sq m', 'factor': 0.092903},
    'sq mile': {'base': 'sq m', 'factor': 2589988.11},

    # Length units -> meters
    'km': {'base': 'm', 'factor': 1000},
    'kilometer': {'base': 'm', 'factor': 1000},
    'kilometers': {'base': 'm', 'factor': 1000},
    'm': {'base': 'm', 'factor': 1},
    'meter': {'base': 'm', 'factor': 1},
    'meters': {'base': 'm', 'factor': 1},
    'cm': {'base': 'm', 'factor': 0.01},
    'mm': {'base': 'm', 'factor': 0.001},
    'ft': {'base': 'm', 'factor': 0.3048},
    'feet': {'base': 'm', 'factor': 0.3048},
    'mile': {'base': 'm', 'factor': 1609.34},

    # Volume units -> liters
    'ML': {'base': 'L', 'factor': 1000000},
    'Megalitre': {'base': 'L', 'factor': 1000000},
    'kL': {'base': 'L', 'factor': 1000},
    'KL': {'base': 'L', 'factor': 1000},
    'L': {'base': 'L', 'factor': 1},
    'litre': {'base': 'L', 'factor': 1},
    'litres': {'base': 'L', 'factor': 1},
    'm³': {'base': 'L', 'factor': 1000},
    'm3': {'base': 'L', 'factor': 1000},
    'cubic meter': {'base': 'L', 'factor': 1000},
    'cubic metres': {'base': 'L', 'factor': 1000},
    'gal': {'base': 'L', 'factor': 3.78541},
    'gallon': {'base': 'L', 'factor': 3.78541},

    # Mass/emissions -> tonnes
    'tCO2e': {'base': 't', 'factor': 1},
    't': {'base': 't', 'factor': 1},
    'tonne': {'base': 't', 'factor': 1},
    'tonnes': {'base': 't', 'factor': 1},
    'kt': {'base': 't', 'factor': 1000},
    'Mt': {'base': 't', 'factor': 1000000},
    'kg': {'base': 't', 'factor': 0.001},
    'g': {'base': 't', 'factor': 0.000001},
    'lb': {'base': 't', 'factor': 0.000453592},

    # Water / air concentration units
    'mg/L': {'base': 'mg/L', 'factor': 1},
    'µg/L': {'base': 'mg/L', 'factor': 0.001},
    'ug/L': {'base': 'mg/L', 'factor': 0.001},
    'ng/L': {'base': 'mg/L', 'factor': 0.000001},
    'µg/m³': {'base': 'µg/m³', 'factor': 1},
    'ug/m3': {'base': 'µg/m³', 'factor': 1},
    'mg/m³': {'base': 'µg/m³', 'factor': 1000},
    'ppm': {'base': 'mg/L', 'factor': 1},     # water (approx.)
    'ppb': {'base': 'mg/L', 'factor': 0.001},

    # Soil / sediment concentration -> mg/kg
    'mg/kg': {'base': 'mg/kg', 'factor': 1},
    'mg/kg dw': {'base': 'mg/kg', 'factor': 1},
    'mg/kg-dw': {'base': 'mg/kg', 'factor': 1},
    'µg/kg': {'base': 'mg/kg', 'factor': 0.001},
    'ug/kg': {'base': 'mg/kg', 'factor': 0.001},
    'ng/kg': {'base': 'mg/kg', 'factor': 0.000001},

    # Flow -> L/s
    'L/s': {'base': 'L/s', 'factor': 1},
    'Lps': {'base': 'L/s', 'factor': 1},
    'm³/s': {'base': 'L/s', 'factor': 1000},
    'm3/s': {'base': 'L/s', 'factor': 1000},
    'ML/d': {'base': 'L/s', 'factor': 11.574},
    'L/min': {'base': 'L/s', 'factor': 1/60},
    'L/hr': {'base': 'L/s', 'factor': 1/3600},

    # Hydrology – rainfall / intensity
    'mm/d': {'base': 'mm/d', 'factor': 1},
    'mm/day': {'base': 'mm/d', 'factor': 1},
    'mm/yr': {'base': 'mm/d', 'factor': 1/365},
    'mm/year': {'base': 'mm/d', 'factor': 1/365},

    # Power
    'MW': {'base': 'MW', 'factor': 1},
    'GW': {'base': 'MW', 'factor': 1000},
    'kW': {'base': 'MW', 'factor': 0.001},
    'W': {'base': 'MW', 'factor': 0.000001},

    # Energy -> MJ
    'GJ': {'base': 'MJ', 'factor': 1000},
    'MJ': {'base': 'MJ', 'factor': 1},
    'kJ': {'base': 'MJ', 'factor': 0.001},
    'MWh': {'base': 'MJ', 'factor': 3600},
    'kWh': {'base': 'MJ', 'factor': 3.6},

    # People
    'people': {'base': 'people', 'factor': 1},
    'persons': {'base': 'people', 'factor': 1},
    'workers': {'base': 'people', 'factor': 1},
    'employees': {'base': 'people', 'factor': 1},

    # Biodiversity – counts
    'species': {'base': 'species', 'factor': 1},
    'species_count': {'base': 'species', 'factor': 1},
    'individuals': {'base': 'individuals', 'factor': 1},
    'trees': {'base': 'individuals', 'factor': 1},
    'birds': {'base': 'individuals', 'factor': 1},
    'fauna': {'base': 'individuals', 'factor': 1},
    'flora': {'base': 'individuals', 'factor': 1},
}


# Parameter context patterns for like-for-like comparison
PARAMETER_CONTEXTS = {
    'study_area': {
        'patterns': [r'study\s+area', r'project\s+area', r'assessment\s+area'],
        'valid_units': ['sq m', 'ha', 'km²', 'hectares'],
    },
    'disturbance_area': {
        'patterns': [r'disturbance\s+area', r'clearing\s+area', r'footprint', r'impact\s+area'],
        'valid_units': ['sq m', 'ha', 'km²', 'hectares'],
    },
    'workforce': {
        'patterns': [r'workforce', r'employees?', r'workers?', r'personnel', r'staff'],
        'valid_units': ['people', 'persons', 'workers', 'employees', ''],
    },
    'water_consumption': {
        'patterns': [r'water\s+(?:consumption|use|demand|requirement)', r'water\s+supply'],
        'valid_units': ['L', 'ML', 'kL', 'm³'],
    },
    'power_capacity': {
        'patterns': [r'installed\s+capacity', r'generation\s+capacity', r'power\s+capacity'],
        'valid_units': ['MW', 'GW', 'kW'],
    },
    'transmission_length': {
        'patterns': [r'transmission\s+line', r'power\s+line', r'cable\s+length'],
        'valid_units': ['m', 'km'],
    },
}

# Domain to Project Summary section mapping
DOMAIN_TO_SECTION = {
    'Project Overview': ['executive_summary', 'project_description', 'introduction'],
    'Environmental & Social Baseline': ['baseline_conditions'],
    'Major Anticipated Impacts': ['environmental_and_social_impact_assessment'],
    'Mitigation & Management Measures': ['mitigation_and_enhancement_measures', 'environmental_and_social_management_plan_esmp'],
    'Residual Risks & Compliance': ['conclusion_and_recommendations', 'PS1', 'PS2', 'PS3', 'PS4', 'PS5', 'PS6', 'PS7', 'PS8']
}

# Gap analysis expected items
GAP_CHECKS = {
    "Project Description": {
        "Location Coordinates": r'\d+°\s*\d+\'\s*\d*"?\s*[NS].*?\d+°\s*\d+\'\s*\d*"?\s*[EW]',
        "Project Area": r'\b\d+(?:,\d{3})*(?:\.\d+)?\s*(?:ha|hectares?|km²|km2)\b',
        "Workforce Numbers": r'(?:workforce|employees?|workers?)\s*(?:of\s*)?\d+[\d,]*',
        "Water Consumption": r'water\s+(?:consumption|use|demand|requirement)[^.]*?\d+[^.]*',
        "Power Consumption": r'(?:power|electricity)\s+(?:consumption|use|demand|requirement)[^.]*?\d+[^.]*',
    }
}


# =============================================================================
# DATA LOADING FUNCTIONS
# =============================================================================

def load_facts_json(path: Path) -> Dict:
    """Load facts JSON file."""
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Warning: Facts file not found: {path}")
        return {}
    except json.JSONDecodeError as e:
        print(f"Warning: Failed to parse facts JSON: {e}")
        return {}


def load_meta_json(path: Path) -> Dict:
    """Load metadata JSON file."""
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Warning: Meta file not found: {path}")
        return {}
    except json.JSONDecodeError as e:
        print(f"Warning: Failed to parse meta JSON: {e}")
        return {}


def load_chunks_jsonl(path: Path, sample_size: int = 10) -> List[Dict]:
    """Load a sample of chunks from JSONL file."""
    chunks = []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            for i, line in enumerate(f):
                if i >= sample_size:
                    break
                line = line.strip()
                if line:
                    try:
                        chunks.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue
    except FileNotFoundError:
        print(f"Warning: Chunks file not found: {path}")
    return chunks


def load_inputs(facts_path: Path, meta_path: Path, chunks_path: Path) -> Tuple[Dict, Dict, List]:
    """Load all input files."""
    print("Loading input files...")
    facts = load_facts_json(facts_path)
    print(f"  Loaded facts: {len(facts.get('sections', {}))} sections")

    meta = load_meta_json(meta_path)
    print(f"  Loaded meta: {meta.get('document', {}).get('original_filename', 'Unknown')}")

    chunks = load_chunks_jsonl(chunks_path)
    print(f"  Loaded chunks: {len(chunks)} sample chunks")

    return facts, meta, chunks


# =============================================================================
# DATA PROCESSING FUNCTIONS
# =============================================================================

def extract_numeric_values(text: str) -> List[Tuple[float, str, str]]:
    """
    Extract numeric values with units from text.

    Returns:
        List of tuples: (value, unit, raw_match)
    """
    results = []

    # Pattern for numbers with optional units
    pattern = r'(\d+(?:,\d{3})*(?:\.\d+)?)\s*(ha|hectares?|km²|km2|m²|m2|MW|GW|kW|km|m|ML|kL|L|m³|m3|people|persons|workers|employees|tonnes?|t)?'

    for match in re.finditer(pattern, text, re.IGNORECASE):
        value_str = match.group(1).replace(',', '')
        try:
            value = float(value_str)
            unit = match.group(2) or ''
            raw = match.group(0)
            results.append((value, unit.lower() if unit else '', raw))
        except ValueError:
            continue

    return results


def normalize_unit(unit: str) -> Tuple[str, float]:
    """
    Normalize a unit to its base unit.

    Returns:
        Tuple of (base_unit, conversion_factor)
    """
    unit_lower = unit.lower().strip()
    if unit_lower in UNIT_CONVERSIONS:
        conv = UNIT_CONVERSIONS[unit_lower]
        return conv['base'], conv['factor']
    return unit, 1.0


def get_all_facts_text(facts: Dict) -> List[Dict]:
    """
    Extract all fact text values with metadata.

    Returns:
        List of dicts with keys: text, section, page_start, page_end, domain, field
    """
    all_facts = []
    sections = facts.get('sections', {})

    for section_name, section_data in sections.items():
        page_start = section_data.get('page_start', 0)
        page_end = section_data.get('page_end', 0)
        extracted_facts = section_data.get('extracted_facts', {})

        for domain, domain_facts in extracted_facts.items():
            if isinstance(domain_facts, dict):
                for field, value in domain_facts.items():
                    if value and isinstance(value, str) and value.strip():
                        all_facts.append({
                            'text': value,
                            'section': section_name,
                            'page_start': page_start,
                            'page_end': page_end,
                            'domain': domain,
                            'field': field,
                        })

    return all_facts


def build_project_summary(facts: Dict) -> Dict[str, str]:
    """
    Build project summary content for the 5 main sections.

    Returns:
        Dict mapping section title to content string
    """
    summary = {}
    sections = facts.get('sections', {})

    for summary_section, domains in DOMAIN_TO_SECTION.items():
        bullets = []

        for section_name, section_data in sections.items():
            extracted_facts = section_data.get('extracted_facts', {})

            for domain in domains:
                if domain in extracted_facts:
                    domain_facts = extracted_facts[domain]
                    if isinstance(domain_facts, dict):
                        for field, value in domain_facts.items():
                            if value and isinstance(value, str) and value.strip():
                                # Clean up the value
                                clean_value = value.strip()
                                if len(clean_value) > 500:
                                    clean_value = clean_value[:500] + "..."
                                bullets.append(f"• {clean_value}")

        if bullets:
            # Deduplicate and limit bullets
            unique_bullets = list(dict.fromkeys(bullets))[:10]
            summary[summary_section] = "\n".join(unique_bullets)
        else:
            summary[summary_section] = "No information extracted"

    return summary


def build_fact_categories(facts: Dict) -> List[Dict]:
    """
    Build fact categories from sections.

    Returns:
        List of dicts with keys: category, count, samples
    """
    categories = []
    sections = facts.get('sections', {})

    section_counts = []
    for section_name, section_data in sections.items():
        chunk_count = section_data.get('chunk_count', 1)
        section_counts.append({
            'category': section_name,
            'count': chunk_count,
            'samples': section_name[:50]
        })

    # Sort by count descending
    section_counts.sort(key=lambda x: -x['count'])

    return section_counts


def check_consistency(facts: Dict) -> List[Dict]:
    """
    Check for consistency issues in numeric values.

    Returns:
        List of consistency issue dicts
    """
    issues = []
    all_facts = get_all_facts_text(facts)

    # Group numeric values by parameter context
    context_values = defaultdict(list)

    for fact in all_facts:
        text = fact['text']

        # Check which parameter context this text belongs to
        for context_name, context_info in PARAMETER_CONTEXTS.items():
            for pattern in context_info['patterns']:
                if re.search(pattern, text, re.IGNORECASE):
                    # Extract numeric values
                    numeric_values = extract_numeric_values(text)
                    for value, unit, raw in numeric_values:
                        if unit.lower() in [u.lower() for u in context_info['valid_units']] or not unit:
                            context_values[context_name].append({
                                'value': value,
                                'unit': unit,
                                'raw': raw,
                                'page': fact['page_start'],
                                'section': fact['section'],
                            })
                    break

    # Check for inconsistencies within each context
    for context_name, values in context_values.items():
        if len(values) < 2:
            continue

        # Normalize all values to base unit
        normalized = []
        for v in values:
            base_unit, factor = normalize_unit(v['unit'])
            normalized.append({
                **v,
                'normalized_value': v['value'] * factor,
                'base_unit': base_unit,
            })

        # Group by base unit
        by_base_unit = defaultdict(list)
        for n in normalized:
            by_base_unit[n['base_unit']].append(n)

        for base_unit, group in by_base_unit.items():
            if len(group) < 2:
                continue

            norm_values = [g['normalized_value'] for g in group]
            min_val = min(norm_values)
            max_val = max(norm_values)

            if min_val > 0:
                diff_percent = ((max_val - min_val) / min_val) * 100
            else:
                diff_percent = 0

            if diff_percent > 5:  # Threshold for significance
                severity = 'high' if diff_percent > 20 else 'medium'

                issues.append({
                    'severity': severity,
                    'context': context_name.replace('_', ' ').title(),
                    'values': [f"{g['raw']} [p.{g['page']}]" for g in group],
                    'normalized_values': [g['normalized_value'] for g in group],
                    'base_unit': base_unit,
                    'diff_percent': round(diff_percent, 1),
                    'message': f"{context_name.replace('_', ' ').title()} reported with {len(group)} different values across the document.",
                })

    return issues


def check_unit_standardization(facts: Dict) -> List[Dict]:
    """
    Check for unit standardization issues.

    Returns:
        List of unit issue dicts
    """
    issues = []
    all_facts = get_all_facts_text(facts)

    # Group by parameter context and track units used
    context_units = defaultdict(lambda: {'units': set(), 'examples': []})

    for fact in all_facts:
        text = fact['text']

        for context_name, context_info in PARAMETER_CONTEXTS.items():
            for pattern in context_info['patterns']:
                if re.search(pattern, text, re.IGNORECASE):
                    numeric_values = extract_numeric_values(text)
                    for value, unit, raw in numeric_values:
                        if unit:
                            context_units[context_name]['units'].add(unit.lower())
                            if len(context_units[context_name]['examples']) < 3:
                                context_units[context_name]['examples'].append({
                                    'value': value,
                                    'unit': unit,
                                    'page': fact['page_start'],
                                })
                    break

    # Report contexts with multiple units
    for context_name, data in context_units.items():
        if len(data['units']) > 1:
            base_unit, _ = normalize_unit(list(data['units'])[0])
            issues.append({
                'context': context_name.replace('_', ' ').title(),
                'units_used': list(data['units']),
                'examples': data['examples'],
                'base_unit': base_unit,
                'recommendation': f"Standardize to {base_unit} for consistency",
            })

    return issues


def check_thresholds(meta: Dict, facts: Dict) -> List[Dict]:
    """
    Check threshold compliance from tables in metadata.

    Returns:
        List of threshold check dicts
    """
    results = []
    tables = meta.get('tables', [])

    # Common environmental parameters and their typical thresholds
    threshold_patterns = {
        'pH': {'pattern': r'\bpH\b.*?(\d+(?:\.\d+)?)', 'min': 6.0, 'max': 9.0, 'unit': ''},
        'BOD': {'pattern': r'BOD.*?(\d+(?:\.\d+)?)\s*mg/[Ll]', 'max': 50, 'unit': 'mg/L'},
        'COD': {'pattern': r'COD.*?(\d+(?:\.\d+)?)\s*mg/[Ll]', 'max': 100, 'unit': 'mg/L'},
        'TSS': {'pattern': r'(?:TSS|suspended\s+solids).*?(\d+(?:\.\d+)?)\s*mg/[Ll]', 'max': 100, 'unit': 'mg/L'},
        'PM10': {'pattern': r'PM\s*10.*?(\d+(?:\.\d+)?)\s*(?:µg|ug)/m', 'max': 50, 'unit': 'µg/m³'},
        'PM2.5': {'pattern': r'PM\s*2\.?5.*?(\d+(?:\.\d+)?)\s*(?:µg|ug)/m', 'max': 25, 'unit': 'µg/m³'},
        'Noise': {'pattern': r'(\d+(?:\.\d+)?)\s*dB\s*\(?A\)?', 'max': 70, 'unit': 'dB(A)'},
    }

    # Scan tables for threshold data
    for table in tables:
        content = table.get('content', '')
        page = table.get('page', 0)

        for param_name, config in threshold_patterns.items():
            matches = re.findall(config['pattern'], content, re.IGNORECASE)
            for match in matches:
                try:
                    value = float(match)
                    threshold = config.get('max', config.get('min', 0))

                    if 'max' in config:
                        status = 'Exceeds' if value > config['max'] else 'Within'
                    elif 'min' in config:
                        status = 'Exceeds' if value < config['min'] else 'Within'
                    else:
                        status = 'Unknown'

                    results.append({
                        'parameter': param_name,
                        'category': 'Environmental',
                        'value': value,
                        'threshold': threshold,
                        'unit': config['unit'],
                        'status': status,
                        'page': page,
                        'source': 'Table extraction',
                    })
                except ValueError:
                    continue

    return results


def analyze_gaps(facts: Dict) -> List[Dict]:
    """
    Analyze gaps in expected content.

    Returns:
        List of gap analysis dicts
    """
    results = []
    all_facts = get_all_facts_text(facts)

    # Combine all text for searching
    all_text = " ".join([f['text'] for f in all_facts])

    for section, items in GAP_CHECKS.items():
        for item_name, pattern in items.items():
            match = re.search(pattern, all_text, re.IGNORECASE)

            if match:
                # Find which fact contains this match
                content_found = match.group(0)[:200]
                pages = []
                for fact in all_facts:
                    if match.group(0) in fact['text']:
                        pages.append(str(fact['page_start']))
                        break

                results.append({
                    'section': section,
                    'subsection': item_name,
                    'status': 'PRESENT',
                    'content_found': content_found,
                    'pages': ', '.join(pages) if pages else '',
                })
            else:
                results.append({
                    'section': section,
                    'subsection': item_name,
                    'status': 'MISSING',
                    'content_found': '',
                    'pages': '',
                })

    return results


# =============================================================================
# EXCEL GENERATION FUNCTIONS
# =============================================================================

def apply_header_style(ws, row: int, num_cols: int):
    """Apply header styling to a row."""
    header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font


def build_summary_sheet(ws, facts: Dict, meta: Dict) -> None:
    """Build the Summary sheet."""
    doc_info = meta.get('document', {})
    stats = meta.get('statistics', {})

    # Title
    ws.cell(1, 1, "ESIA Review Analysis Summary").font = Font(size=16, bold=True)
    ws.cell(1, 1).alignment = Alignment(horizontal='left')

    row = 3
    summary_data = [
        ("Document", doc_info.get('original_filename', 'Unknown')),
        ("Processed Filename", doc_info.get('processed_filename', 'Unknown')),
        ("Total Chunks", facts.get('total_chunks', stats.get('total_chunks', 0))),
        ("Total Pages", doc_info.get('total_pages', 0)),
        ("Sections Processed", facts.get('sections_processed', 0)),
        ("Sections with Facts", facts.get('sections_with_facts', 0)),
        ("Analysis Date", datetime.now().isoformat()),
    ]

    for key, value in summary_data:
        ws.cell(row, 1, key).font = Font(bold=True)
        ws.cell(row, 2, str(value))
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60


def build_project_summary_sheet(ws, project_summary: Dict) -> None:
    """Build the Project Summary sheet."""
    headers = ["Section", "Content"]
    ws.append(headers)
    apply_header_style(ws, 1, len(headers))

    section_order = [
        "Project Overview",
        "Environmental & Social Baseline",
        "Major Anticipated Impacts",
        "Mitigation & Management Measures",
        "Residual Risks & Compliance"
    ]

    for section in section_order:
        content = project_summary.get(section, "No information extracted")
        ws.append([section, content])

    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 100

    # Wrap text in content column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')


def build_fact_categories_sheet(ws, categories: List[Dict]) -> None:
    """Build the Fact Categories sheet."""
    headers = ["Category", "Count", "Sample Sections"]
    ws.append(headers)
    apply_header_style(ws, 1, len(headers))

    for cat in categories:
        ws.append([cat['category'], cat['count'], cat['samples']])

    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50


def build_consistency_issues_sheet(ws, issues: List[Dict]) -> None:
    """Build the Consistency Issues sheet."""
    headers = ["Severity", "Parameter Context", "Values Found", "Normalized (base unit)", "Difference %", "Details"]
    ws.append(headers)
    apply_header_style(ws, 1, len(headers))

    red_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    red_font = Font(color="721c24")
    orange_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    orange_font = Font(color="856404")

    for issue in issues:
        row_data = [
            issue.get('severity', '').upper(),
            issue.get('context', ''),
            "; ".join(issue.get('values', [])),
            f"{issue.get('normalized_values', [])} ({issue.get('base_unit', '')})",
            f"{issue.get('diff_percent', 0)}%",
            issue.get('message', ''),
        ]
        ws.append(row_data)

        # Apply severity styling
        row_idx = ws.max_row
        if issue.get('severity') == 'high':
            for cell in ws[row_idx]:
                cell.fill = red_fill
                cell.font = red_font
        elif issue.get('severity') == 'medium':
            for cell in ws[row_idx]:
                cell.fill = orange_fill
                cell.font = orange_font

    # Adjust column widths
    for i, width in enumerate([12, 20, 40, 30, 15, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_unit_standardization_sheet(ws, unit_issues: List[Dict]) -> None:
    """Build the Unit Standardization sheet."""
    headers = ["Parameter Context", "Units Used", "Examples", "Recommendation"]
    ws.append(headers)
    apply_header_style(ws, 1, len(headers))

    for issue in unit_issues:
        examples_str = "; ".join([
            f"{e['value']} {e['unit']} (p.{e['page']})"
            for e in issue.get('examples', [])
        ])
        row_data = [
            issue.get('context', ''),
            ", ".join(issue.get('units_used', [])),
            examples_str,
            issue.get('recommendation', ''),
        ]
        ws.append(row_data)

    # Adjust column widths
    for i, width in enumerate([25, 20, 50, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_threshold_compliance_sheet(ws, threshold_checks: List[Dict]) -> None:
    """Build the Threshold Compliance sheet."""
    headers = ["Parameter", "Category", "Value", "Threshold", "Unit", "Status", "Page", "Source"]
    ws.append(headers)
    apply_header_style(ws, 1, len(headers))

    red_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")

    for check in threshold_checks:
        row_data = [
            check.get('parameter', ''),
            check.get('category', ''),
            check.get('value', ''),
            check.get('threshold', ''),
            check.get('unit', ''),
            check.get('status', ''),
            check.get('page', ''),
            check.get('source', ''),
        ]
        ws.append(row_data)

        # Apply status styling
        row_idx = ws.max_row
        if check.get('status') == 'Exceeds':
            for cell in ws[row_idx]:
                cell.fill = red_fill
        elif check.get('status') == 'Within':
            for cell in ws[row_idx]:
                cell.fill = green_fill

    # Adjust column widths
    for i, width in enumerate([15, 15, 10, 10, 10, 10, 8, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_gap_analysis_sheet(ws, gaps: List[Dict]) -> None:
    """Build the Gap Analysis sheet."""
    headers = ["Section", "Sub-section", "Status", "Content Found", "Page(s)"]
    ws.append(headers)
    apply_header_style(ws, 1, len(headers))

    green_font = Font(color="155724")
    red_font = Font(color="721c24", bold=True)

    for gap in gaps:
        row_data = [
            gap.get('section', ''),
            gap.get('subsection', ''),
            gap.get('status', ''),
            gap.get('content_found', '')[:200] if gap.get('content_found') else '',
            gap.get('pages', ''),
        ]
        ws.append(row_data)

        # Apply status styling
        row_idx = ws.max_row
        status_cell = ws.cell(row=row_idx, column=3)
        if gap.get('status') == 'PRESENT':
            status_cell.font = green_font
        else:
            status_cell.font = red_font

    # Adjust column widths
    for i, width in enumerate([25, 25, 12, 60, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_excel(output_path: Path, data: Dict) -> None:
    """Generate the complete Excel workbook."""
    print("Generating Excel workbook...")

    wb = Workbook()

    # Summary sheet (use active sheet)
    ws_summary = wb.active
    ws_summary.title = "Summary"
    build_summary_sheet(ws_summary, data['facts'], data['meta'])
    print("  Built Summary sheet")

    # Project Summary sheet
    ws_project = wb.create_sheet("Project Summary")
    build_project_summary_sheet(ws_project, data['project_summary'])
    print("  Built Project Summary sheet")

    # Fact Categories sheet
    ws_categories = wb.create_sheet("Fact Categories")
    build_fact_categories_sheet(ws_categories, data['categories'])
    print("  Built Fact Categories sheet")

    # Consistency Issues sheet
    ws_consistency = wb.create_sheet("Consistency Issues")
    build_consistency_issues_sheet(ws_consistency, data['consistency_issues'])
    print("  Built Consistency Issues sheet")

    # Unit Standardization sheet
    ws_units = wb.create_sheet("Unit Standardization")
    build_unit_standardization_sheet(ws_units, data['unit_issues'])
    print("  Built Unit Standardization sheet")

    # Threshold Compliance sheet
    ws_thresholds = wb.create_sheet("Threshold Compliance")
    build_threshold_compliance_sheet(ws_thresholds, data['threshold_checks'])
    print("  Built Threshold Compliance sheet")

    # Gap Analysis sheet
    ws_gaps = wb.create_sheet("Gap Analysis")
    build_gap_analysis_sheet(ws_gaps, data['gaps'])
    print("  Built Gap Analysis sheet")

    # Save workbook
    wb.save(output_path)
    print(f"  Saved Excel file: {output_path}")


# =============================================================================
# HTML GENERATION FUNCTIONS
# =============================================================================

def escape_html(text: str) -> str:
    """Escape HTML special characters."""
    if not text:
        return ""
    return (text
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&#39;"))


def build_html_header(facts: Dict, meta: Dict) -> str:
    """Build HTML header section."""
    doc_info = meta.get('document', {})
    stats = meta.get('statistics', {})

    return f"""
    <header>
        <h1>ESIA Review Factsheet</h1>
        <p class="subtitle">Document: {escape_html(doc_info.get('original_filename', 'Unknown'))}</p>

        <table class="summary-table">
            <tr><th>Metric</th><th>Value</th></tr>
            <tr><td>Total Pages</td><td>{doc_info.get('total_pages', 0)}</td></tr>
            <tr><td>Total Chunks</td><td>{facts.get('total_chunks', stats.get('total_chunks', 0))}</td></tr>
            <tr><td>Sections with Facts</td><td>{facts.get('sections_with_facts', 0)}</td></tr>
            <tr><td>Analysis Date</td><td>{datetime.now().strftime('%Y-%m-%d %H:%M')}</td></tr>
        </table>
    </header>
    """


def build_html_project_summary(project_summary: Dict) -> str:
    """Build HTML project summary section."""
    html = '<section id="project-summary">\n<h2>Project Summary</h2>\n'

    section_order = [
        "Project Overview",
        "Environmental & Social Baseline",
        "Major Anticipated Impacts",
        "Mitigation & Management Measures",
        "Residual Risks & Compliance"
    ]

    for section in section_order:
        content = project_summary.get(section, "No information extracted")
        html += f'<div class="summary-section">\n<h3>{escape_html(section)}</h3>\n'

        if content == "No information extracted":
            html += f'<p class="no-data">{content}</p>\n'
        else:
            # Convert bullet points to list
            bullets = content.split('\n')
            html += '<ul>\n'
            for bullet in bullets:
                if bullet.strip():
                    clean_bullet = bullet.strip().lstrip('•').strip()
                    html += f'<li>{escape_html(clean_bullet)}</li>\n'
            html += '</ul>\n'

        html += '</div>\n'

    html += '</section>\n'
    return html


def build_html_fact_categories(categories: List[Dict]) -> str:
    """Build HTML fact categories section."""
    html = '<section id="fact-categories">\n<h2>Fact Categories</h2>\n'
    html += '<table>\n<tr><th>Category</th><th>Count</th><th>Sample Sections</th></tr>\n'

    for cat in categories[:20]:  # Limit to top 20
        html += f'<tr><td>{escape_html(cat["category"])}</td><td>{cat["count"]}</td><td>{escape_html(cat["samples"])}</td></tr>\n'

    html += '</table>\n</section>\n'
    return html


def build_html_consistency_issues(issues: List[Dict]) -> str:
    """Build HTML consistency issues section."""
    if not issues:
        return '<section id="consistency-issues">\n<h2>Consistency Issues</h2>\n<p class="no-data">No consistency issues identified.</p>\n</section>\n'

    html = '<section id="consistency-issues">\n<h2>Consistency Issues</h2>\n'
    html += '<table>\n<tr><th>Severity</th><th>Parameter Context</th><th>Values Found</th><th>Normalized</th><th>Difference %</th><th>Details</th></tr>\n'

    for issue in issues:
        severity_class = 'high' if issue.get('severity') == 'high' else 'medium'
        values_str = "; ".join(issue.get('values', []))
        norm_str = f"{issue.get('normalized_values', [])} ({issue.get('base_unit', '')})"

        html += f'<tr class="{severity_class}">'
        html += f'<td>{escape_html(issue.get("severity", "").upper())}</td>'
        html += f'<td>{escape_html(issue.get("context", ""))}</td>'
        html += f'<td>{escape_html(values_str)}</td>'
        html += f'<td>{escape_html(norm_str)}</td>'
        html += f'<td>{issue.get("diff_percent", 0)}%</td>'
        html += f'<td>{escape_html(issue.get("message", ""))}</td>'
        html += '</tr>\n'

    html += '</table>\n</section>\n'
    return html


def build_html_unit_standardization(unit_issues: List[Dict]) -> str:
    """Build HTML unit standardization section."""
    if not unit_issues:
        return '<section id="unit-standardization">\n<h2>Unit Standardization</h2>\n<p class="no-data">All units are standardized.</p>\n</section>\n'

    html = '<section id="unit-standardization">\n<h2>Unit Standardization</h2>\n'
    html += '<table>\n<tr><th>Parameter Context</th><th>Units Used</th><th>Examples</th><th>Recommendation</th></tr>\n'

    for issue in unit_issues:
        examples_str = "; ".join([f"{e['value']} {e['unit']} (p.{e['page']})" for e in issue.get('examples', [])])
        units_str = ", ".join(issue.get('units_used', []))

        html += f'<tr>'
        html += f'<td>{escape_html(issue.get("context", ""))}</td>'
        html += f'<td>{escape_html(units_str)}</td>'
        html += f'<td>{escape_html(examples_str)}</td>'
        html += f'<td>{escape_html(issue.get("recommendation", ""))}</td>'
        html += '</tr>\n'

    html += '</table>\n</section>\n'
    return html


def build_html_threshold_compliance(threshold_checks: List[Dict]) -> str:
    """Build HTML threshold compliance section."""
    if not threshold_checks:
        return '<section id="threshold-compliance">\n<h2>Threshold Compliance</h2>\n<p class="no-data">No threshold data available for comparison.</p>\n</section>\n'

    html = '<section id="threshold-compliance">\n<h2>Threshold Compliance</h2>\n'
    html += '<table>\n<tr><th>Parameter</th><th>Category</th><th>Value</th><th>Threshold</th><th>Unit</th><th>Status</th><th>Page</th><th>Source</th></tr>\n'

    for check in threshold_checks:
        status_class = 'exceeds' if check.get('status') == 'Exceeds' else 'within'

        html += f'<tr class="{status_class}">'
        html += f'<td>{escape_html(check.get("parameter", ""))}</td>'
        html += f'<td>{escape_html(check.get("category", ""))}</td>'
        html += f'<td>{check.get("value", "")}</td>'
        html += f'<td>{check.get("threshold", "")}</td>'
        html += f'<td>{escape_html(check.get("unit", ""))}</td>'
        html += f'<td>{escape_html(check.get("status", ""))}</td>'
        html += f'<td>{check.get("page", "")}</td>'
        html += f'<td>{escape_html(check.get("source", ""))}</td>'
        html += '</tr>\n'

    html += '</table>\n</section>\n'
    return html


def build_html_gap_analysis(gaps: List[Dict]) -> str:
    """Build HTML gap analysis section."""
    html = '<section id="gap-analysis">\n<h2>Gap Analysis</h2>\n'
    html += '<table>\n<tr><th>Section</th><th>Sub-section</th><th>Status</th><th>Content Found</th><th>Page(s)</th></tr>\n'

    for gap in gaps:
        status_class = 'present' if gap.get('status') == 'PRESENT' else 'missing'
        content = gap.get('content_found', '')[:150] + '...' if len(gap.get('content_found', '')) > 150 else gap.get('content_found', '')

        html += f'<tr>'
        html += f'<td>{escape_html(gap.get("section", ""))}</td>'
        html += f'<td>{escape_html(gap.get("subsection", ""))}</td>'
        html += f'<td class="{status_class}">{escape_html(gap.get("status", ""))}</td>'
        html += f'<td>{escape_html(content)}</td>'
        html += f'<td>{escape_html(gap.get("pages", ""))}</td>'
        html += '</tr>\n'

    html += '</table>\n</section>\n'
    return html


def build_html_factsheet(output_path: Path, data: Dict) -> None:
    """Generate the complete HTML factsheet."""
    print("Generating HTML factsheet...")

    doc_name = data['meta'].get('document', {}).get('original_filename', 'ESIA Document')

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ESIA Review - {escape_html(doc_name)}</title>
    <style>
        :root {{
            --primary-color: #1f4e79;
            --success-color: #155724;
            --danger-color: #721c24;
            --warning-color: #856404;
            --light-bg: #f8f9fa;
            --border-color: #dee2e6;
        }}

        * {{
            box-sizing: border-box;
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            line-height: 1.6;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background: #fff;
            color: #333;
        }}

        h1 {{
            color: var(--primary-color);
            border-bottom: 3px solid var(--primary-color);
            padding-bottom: 10px;
        }}

        h2 {{
            color: var(--primary-color);
            margin-top: 2em;
            border-bottom: 1px solid var(--border-color);
            padding-bottom: 5px;
        }}

        h3 {{
            color: #444;
            margin-top: 1.5em;
        }}

        .subtitle {{
            color: #666;
            font-size: 1.1em;
            margin-bottom: 1em;
        }}

        table {{
            border-collapse: collapse;
            width: 100%;
            margin: 1em 0;
            font-size: 0.9em;
        }}

        th {{
            background: var(--primary-color);
            color: white;
            padding: 10px 8px;
            text-align: left;
            font-weight: 600;
        }}

        td {{
            border: 1px solid var(--border-color);
            padding: 8px;
            vertical-align: top;
        }}

        tr:nth-child(even) {{
            background: var(--light-bg);
        }}

        .summary-table {{
            width: auto;
            min-width: 400px;
        }}

        .summary-table th {{
            width: 200px;
        }}

        .summary-section {{
            background: var(--light-bg);
            padding: 15px;
            margin: 1em 0;
            border-radius: 5px;
            border-left: 4px solid var(--primary-color);
        }}

        .summary-section h3 {{
            margin-top: 0;
            color: var(--primary-color);
        }}

        .summary-section ul {{
            margin: 0.5em 0;
            padding-left: 1.5em;
        }}

        .summary-section li {{
            margin: 0.3em 0;
        }}

        .no-data {{
            color: #666;
            font-style: italic;
        }}

        .high {{
            background: #f8d7da !important;
            color: var(--danger-color);
        }}

        .medium {{
            background: #fff3cd !important;
            color: var(--warning-color);
        }}

        .present {{
            color: var(--success-color);
            font-weight: bold;
        }}

        .missing {{
            color: var(--danger-color);
            font-weight: bold;
        }}

        .exceeds {{
            background: #f8d7da !important;
        }}

        .within {{
            background: #d4edda !important;
        }}

        @media print {{
            body {{
                max-width: 100%;
                padding: 0;
            }}

            section {{
                page-break-inside: avoid;
            }}
        }}
    </style>
</head>
<body>
"""

    # Add sections
    html += build_html_header(data['facts'], data['meta'])
    print("  Built header section")

    html += build_html_project_summary(data['project_summary'])
    print("  Built project summary section")

    html += build_html_fact_categories(data['categories'])
    print("  Built fact categories section")

    html += build_html_consistency_issues(data['consistency_issues'])
    print("  Built consistency issues section")

    html += build_html_unit_standardization(data['unit_issues'])
    print("  Built unit standardization section")

    html += build_html_threshold_compliance(data['threshold_checks'])
    print("  Built threshold compliance section")

    html += build_html_gap_analysis(data['gaps'])
    print("  Built gap analysis section")

    html += """
</body>
</html>
"""

    # Write to file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"  Saved HTML file: {output_path}")


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def main() -> None:
    """Main entry point for the ESIA factsheet generator."""
    print("=" * 60)
    print("ESIA REVIEW FACTSHEET GENERATOR")
    print("=" * 60)
    print()

    # Load inputs
    facts, meta, chunks = load_inputs(FACTS_PATH, META_PATH, CHUNKS_PATH)

    if not facts:
        print("Error: Could not load facts file. Exiting.")
        return

    print()
    print("Processing data...")

    # Build all data structures
    project_summary = build_project_summary(facts)
    print("  Built project summary")

    categories = build_fact_categories(facts)
    print(f"  Built {len(categories)} fact categories")

    consistency_issues = check_consistency(facts)
    print(f"  Found {len(consistency_issues)} consistency issues")

    unit_issues = check_unit_standardization(facts)
    print(f"  Found {len(unit_issues)} unit standardization issues")

    threshold_checks = check_thresholds(meta, facts)
    print(f"  Found {len(threshold_checks)} threshold checks")

    gaps = analyze_gaps(facts)
    present_count = sum(1 for g in gaps if g['status'] == 'PRESENT')
    print(f"  Gap analysis: {present_count}/{len(gaps)} items present")

    # Prepare data for export
    data = {
        'facts': facts,
        'meta': meta,
        'chunks': chunks,
        'project_summary': project_summary,
        'categories': categories,
        'consistency_issues': consistency_issues,
        'unit_issues': unit_issues,
        'threshold_checks': threshold_checks,
        'gaps': gaps,
    }

    # Determine output paths
    base_name = FACTS_PATH.stem.replace('_facts', '')
    output_dir = FACTS_PATH.parent

    excel_path = output_dir / f"{base_name}_ESIA_review.xlsx"
    html_path = output_dir / f"{base_name}_ESIA_review.html"

    print()

    # Generate outputs
    generate_excel(excel_path, data)
    print()

    build_html_factsheet(html_path, data)

    print()
    print("=" * 60)
    print("GENERATION COMPLETE")
    print("=" * 60)
    print(f"Excel file: {excel_path.absolute()}")
    print(f"HTML file:  {html_path.absolute()}")
    print()


if __name__ == "__main__":
    main()
